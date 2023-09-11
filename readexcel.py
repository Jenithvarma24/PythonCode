import openpyxl


def open_excel_file(file_path):
    try:
        return openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print('Excel file not found. Please provide a valid file path.')
        return None


def get_sheet_by_name(workbook, sheet_name):
    try:
        sheet = workbook[sheet_name]
        return sheet
    except KeyError:
        print(f'Sheet "{sheet_name}" does not exist in the workbook.')
        return None


def read_excel_row(file_path, sheet_name, row_index):
    workbook = open_excel_file(file_path)
    if not workbook:
        return None

    sheet = get_sheet_by_name(workbook, sheet_name)
    if not sheet:
        return None

    try:
        row_data = list(sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True))[0]
        return row_data
    except ValueError:
        print('Invalid input. Please enter a numeric row index.')


def read_excel_column(file_path, sheet_name, column_index):
    workbook = open_excel_file(file_path)
    if not workbook:
        return None

    sheet = get_sheet_by_name(workbook, sheet_name)
    if not sheet:
        return None

    try:
        column_data = [cell[0] for cell in
                       sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True)]
        return column_data
    except ValueError:
        print('Invalid input. Please enter a numeric column index.')


def read_excel_cell(file_path, sheet_name, row_index, column_index):
    workbook = open_excel_file(file_path)
    if not workbook:
        return None

    sheet = get_sheet_by_name(workbook, sheet_name)
    if not sheet:
        return None

    try:
        cell_data = sheet.cell(row=row_index, column=column_index).value
        return cell_data
    except ValueError:
        print('Invalid input. Please enter numeric row and column indices.')


if __name__ == "__main__":
    excel_file_path = r'C:\Users\jenith.ravichandran\PycharmProjects\pythonProject2\ExceptionAndFilehandiling\paln.xlsx'  # Replace with the path to your Excel file (corrected the typo in the file name)
    sheet_name = input('Enter the sheet name you want to read: ')
    row_index = int(input('Enter the desired row index : '))
    column_index = int(input('Enter the desired column index : '))

    row_data = read_excel_row(excel_file_path, sheet_name, row_index)
    column_data = read_excel_column(excel_file_path, sheet_name, column_index)
    cell_data = read_excel_cell(excel_file_path, sheet_name, row_index, column_index)

    if row_data:
        print(f'Row {row_index} data: {row_data}')

    if column_data:
        print(f'Column {column_index} data: {column_data}')

    if cell_data:
        print(f'Cell ({row_index}, {column_index}) data: {cell_data}')
